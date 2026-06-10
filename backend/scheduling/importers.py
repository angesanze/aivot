"""
Import massivo di persone da file Excel (.xlsx) o CSV.

Formato atteso: una persona per riga.
  - colonna 1: nome (obbligatorio)
  - colonna 2: competenze, separate da virgola o punto e virgola (opzionale)

La prima riga può essere un'intestazione (es. "Nome; Competenze"): viene
riconosciuta e saltata. Le righe senza nome vengono scartate e segnalate.
"""
import csv
import io

MAX_ROWS = 2000
MAX_SIZE = 2 * 1024 * 1024  # 2 MB

HEADER_WORDS = {"nome", "name", "persona", "cognome"}


class ImportError_(Exception):
    """Errore d'import mostrabile all'utente."""


def _rows_from_csv(data):
    text = data.decode("utf-8-sig", errors="replace")
    # Delimitatore: virgola o punto e virgola, quale dei due è più usato
    delimiter = ";" if text.count(";") > text.count(",") else ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _rows_from_xlsx(data):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise ImportError_("File Excel non leggibile: salvalo come .xlsx "
                           "e riprova.")
    sheet = wb.active
    return [["" if c is None else str(c) for c in row]
            for row in sheet.iter_rows(values_only=True)]


def parse_people_file(filename, data):
    """[(nome, [skills])] dal file, più la lista dei numeri di riga scartati."""
    if len(data) > MAX_SIZE:
        raise ImportError_("File troppo grande (massimo 2 MB).")

    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        rows = _rows_from_xlsx(data)
    elif name.endswith(".csv") or name.endswith(".txt"):
        rows = _rows_from_csv(data)
    elif name.endswith(".xls"):
        raise ImportError_("Il vecchio formato .xls non è supportato: "
                           "salva il file come .xlsx.")
    else:
        raise ImportError_("Formato non riconosciuto: carica un file "
                           ".xlsx o .csv.")

    if len(rows) > MAX_ROWS + 1:
        raise ImportError_(f"Troppe righe (massimo {MAX_ROWS}).")

    # Intestazione: se la prima cella sembra un titolo di colonna, si salta
    start = 0
    if rows and rows[0] and rows[0][0].strip().lower() in HEADER_WORDS:
        start = 1

    people, skipped = [], []
    for i, row in enumerate(rows[start:], start=start + 1):
        name = (row[0] if row else "").strip()
        if not name:
            if any(str(c).strip() for c in row):
                skipped.append(i)  # riga con dati ma senza nome
            continue  # righe del tutto vuote: ignorate in silenzio
        raw_skills = (row[1] if len(row) > 1 else "") or ""
        skills = [s.strip() for s in raw_skills.replace(";", ",").split(",")
                  if s.strip()]
        people.append((name, skills))

    if not people:
        raise ImportError_("Nessuna persona trovata nel file: serve il nome "
                           "nella prima colonna.")
    return people, skipped
